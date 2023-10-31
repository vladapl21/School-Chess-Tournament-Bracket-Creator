from openpyxl import Workbook, load_workbook
from main import next_round_matches

wb = load_workbook("Chess_Tournament.xlsx")
ws = wb["Round 7"]


def import_round(next_round):
    for row in range(1, 53):
        for col in range(2):
            char = chr(65 + col)
            ws[str(char) + str(row)] = next_round[row-1][col]


import_round(next_round_matches)

wb.save("Chess_Tournament.xlsx")
